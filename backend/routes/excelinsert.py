from flask import Flask, request, jsonify, Blueprint, send_file
from models import db
import zipfile
import tempfile
import random

import re
from models.User import User
from models.Qs import QuizQuestion
from models.Paper import Paper
from openpyxl import load_workbook
import os
import datetime as datetime
excelInsert_bp = Blueprint('excelInsert', __name__)
TEMPLATE_PATH = "Skill matric evaluation SAP.xlsx"

EXPORT_BASE = r"C:\\Users\\User\\Desktop\\ExcelSaves"


def fill_excel_template(data, output_path, optionSelected, template_path='Skill matric evaluation SAP.xlsx'):
    wb = load_workbook(template_path)
    ws = wb.active
    count = 0
    for row in ws.iter_rows():
        for cell in row:
            try:
                key = str(cell.value).strip() if cell.value else None
                # Fill simple fields
                if key and key in data and key != "Questionaries":
                    ws.cell(row=cell.row, column=cell.column + 1).value = data[key]
                # Handle Questionaries block
                if key == "Questionaries":
                    start_row = cell.row + 1
                    col = cell.column
                    for i, q in enumerate(data["Questionaries"]):
                        row_idx = start_row + i

                        # Insert the question
                        ws.cell(row=row_idx, column=col).value = q["question"]

                        # Insert the scores (optional, if present)
                        if "scores" in q:
                            for j, score in enumerate(q["scores"]):
                                ws.cell(row=row_idx, column=col + j + 1).value = score  # B, C, D, E

                        # Insert the correct option (optional, if present)
                        # For the current question index (i)
                        selected = optionSelected[i] if i < len(optionSelected) else 'NA'
                        if selected != 'NA':
                            selectedLetter = chr(65 + int(selected))
                        else:
                            selectedLetter = "Not Attempted"

                        ws.cell(row=row_idx, column=col + 5).value = selectedLetter

                        if 'correct_option' in q:
                            correctLetter = chr(65+q['correct_option'])
                            isCorrect = selectedLetter == correctLetter
                            if isCorrect:
                                count+=2
                            ws.cell(row=row_idx, column=col + 6).value = "True" if isCorrect else "False"
                
                
                if key == "Total Obtained  Marks":
                    ws.cell(row=cell.row, column=cell.column + 1).value = count
                if key == 'Achieved  Percentage':
                    total_questions = len(data["Questionaries"])
                    total_marks = total_questions * 2
                    
                    perc = (count / total_marks) * 100 if total_marks > 0 else 0
                    ws.cell(row=cell.row, column=cell.column + 1).value = round(perc / 100, 4)  # Store as decimal for Excel percent
                    ws.cell(row=cell.row, column=cell.column + 1).number_format = '0.00%'  # Format as percent in Excel


            except Exception as e:
                print(f"Error at cell {cell.coordinate if cell else 'unknown'}: {e}")

    wb.save(output_path)


@excelInsert_bp.route("/manual", methods=["POST"])
def generate_skill_matrix_excels():
    try:
        data = request.get_json()
        optionSelected = data.get('valuesOnly')
        user_id = data.get('user_id')
        paperId = data.get('paperId')
        print(user_id, paperId)
        if not isinstance(optionSelected, list) or not user_id:
            return jsonify({"error": "Missing or invalid 'valuesOnly' or 'userId'"}), 400

        paper = Paper.query.filter_by(paperId=paperId).first()
        paperTitle = paper.title
        user = User.query.filter_by(tclid=user_id).first()
        if not user:
            return jsonify({"error": "User not found"}), 404
        # Replace problematic characters with underscores (spaces, &, /, etc.)
        safe_title = re.sub(r'[^\w\- ]', '_', paperTitle)
        EXPORT_DIR = os.path.join(EXPORT_BASE, safe_title)
        os.makedirs(EXPORT_DIR, exist_ok=True)

        filename = f"{user.name.replace(' ', '_')}_{safe_title}_SkillMatrix.xlsx"
        full_path = os.path.join(EXPORT_DIR, filename)
        
        rev = random.randint(0,99)

        qs = QuizQuestion.query.filter_by(paper_id=paperId).all()
        questions = [{"question": q.qs, "scores": q.options, "correct_option": q.correct_option} for q in qs]
        TMA = len(questions) * 2
        # Calculate percentage score from selected answers
        count = 0
        for i, q in enumerate(questions):
            selected = optionSelected[i] if i < len(optionSelected) else 'NA'
            if selected != 'NA' and int(selected) == q['correct_option']:
                count += 2

        perc = (count / TMA) * 100 if TMA > 0 else 0

        # Determine Current Skill Level
        if perc >= 80:
            CSL = 'Expert'
        elif 60 <= perc < 80:
            CSL = 'Amateur'
        elif 40 <= perc < 60:
            CSL = 'Beginner'
        else:
            CSL = 'Failed'


        paperTitle = paper.title

        evaluator_map = {
            'General Tech Quiz': 'Vijaya Laxmi',
            'E&D': 'Sangeet Raman',
            'Vendor': 'Ravi Deshmukh',
            'HR': 'Pooja Sharma',
            'TPM': 'Rajesh Nair',
            'Purchase': 'Puneet Sharma',
            'Finance': 'Amit Bhagat',
            'IT': 'Sunil Bidla',
            'Production': 'Santosh Phad',
            'Maintenance': 'Suresh Jadhav',
            'Quality': 'Babarao ',
            'Dispatch': 'Deepak Yadav'
        }

        skillMap = {
            'General Tech Quiz': 'Understanding of basic general knowledge, logical reasoning, and current affairs.',
            'E&D': 'Knowledge of Engineering drawings, GD&T, product development cycles, and documentation standards.',
            'Vendor': 'Vendor evaluation, negotiation skills, supply chain communication, and onboarding process.',
            'HR': 'Understanding of recruitment, employee engagement, payroll systems, and HR policies.',
            'TPM': 'Competency in Total Productive Maintenance pillars, autonomous maintenance, and Kaizen implementation.',
            'Purchase': 'Procurement cycle knowledge, inventory control, vendor negotiation, and SAP purchasing module.',
            'Finance': 'Understanding of cost centers, budgeting, financial reporting, taxation, and compliance.',
            'IT': 'Basic Knowledge of SAP modules, Networking, security protocols, system troubleshooting, and IT asset management.',
            'Production': 'Knowledge of production planning, OEE, process optimization, lean principles, and safety standards.',
            'Maintenance': 'Skill in predictive/preventive maintenance, troubleshooting machinery, and TPM practices.',
            'Quality': 'Competency in QC tools, root cause analysis, ISO standards, and continuous improvement practices.',
            'Dispatch': 'Understanding of dispatch planning, logistics coordination, invoice creation, and SAP SD module.'
        }

        skill = skillMap.get(paperTitle, 'General Knowledge')
        evalName = evaluator_map.get(paperTitle, 'Unknown Evaluator')
        docN = f"WI-HRA-{rev}"
        data_to_fill = {
            "Employee Name": user.name,
            "Evaluator Name": evalName,
            "Dept./ Section": paperTitle,
            "Current Skill level": CSL,
            "Rev Date :": datetime.datetime.today().strftime("%Y-%m-%d"),
            "Prepared  By:-": "VijayaLaxmi",
            "Approved By:-": "Avinash Chormale",
            "Questionaries": questions,
            "Total  Marks available": TMA,
            "Rev No." : rev,
            "Doc No. :" : docN,
            "Skills/ Competence" : skill,
        }

        filename = f"{user.name.replace(' ', '_')}_SkillMatrix.xlsx"
        full_path = os.path.join(EXPORT_DIR, filename )
        fill_excel_template(data_to_fill, full_path, optionSelected)

        return jsonify({
            "message": "Excel file generated.",
            "file": filename
        }), 200

    except Exception as e:
        print(e)
        return jsonify({"error": str(e)}), 500


@excelInsert_bp.route('/downloadZip/<int:paper_id>', methods=["GET"])
def downloadZip(paper_id):
    try:
        user = User.query.filter
        if user == 0:
            return jsonify({"error":"No Reports Available"})
        paper = Paper.query.filter_by(paperId=paper_id).first()
        if not paper:
            return jsonify({"error": "Paper not found"}), 400

        safe_title = re.sub(r'[^\w\- ]', '_', paper.title)
        EXPORT_DIR = os.path.join(EXPORT_BASE, safe_title)

        if not os.path.exists(EXPORT_DIR) or not os.listdir(EXPORT_DIR):
            return jsonify({"error": "No Excel files found for this paper."}), 404

        with tempfile.NamedTemporaryFile(delete=False, suffix='.zip') as tmp_zip:
            with zipfile.ZipFile(tmp_zip, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for root, _, files in os.walk(EXPORT_DIR):
                    for file in files:
                        print('Zipping file:', file)
                        file_path = os.path.join(root, file)
                        arcname = os.path.relpath(file_path, EXPORT_DIR)
                        zipf.write(file_path, arcname)
        return send_file(
            tmp_zip.name,
            mimetype='application/zip',
            as_attachment=True,
            download_name=f"{safe_title}_SkillMatrix.zip",
        )

    except Exception as e:
        print("Error in zip:", e)
        return jsonify({'error': 'Download Zip Error'}), 500
