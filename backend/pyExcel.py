from openpyxl import load_workbook

def fill_excel_template(data, output_path, template_path='Skill matric evaluation SAP.xlsx'):
    wb = load_workbook(template_path)
    ws = wb.active

    for row in ws.iter_rows():
        for cell in row:
            try:
                key = str(cell.value).strip() if cell.value else None

                # Fill simple fields
                if key and key in data and key != "Questionaries":
                    ws.cell(row=cell.row, column=cell.column + 1).value = data[key]

                # Handle Questionaries block
                if key == "Questionaries":
                    start_row = cell.row + 1
                    col = cell.column
                    for i, q in enumerate(data["Questionaries"]):
                        ws.cell(row=start_row + i, column=col).value = q["question"]
                        for j, score in enumerate(q["scores"]):
                            ws.cell(row=start_row + i, column=col + j + 1).value = score  # A, B, C, D
            except Exception as e:
                print(f"Error at cell {cell.coordinate if cell else 'unknown'}: {e}")

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    # Each question has associated 4 scores (for A, B, C, D)
    questions = [
        {"question": "What is SAP", "scores": [12, 26, 24, 58]},
        {"question": "Define ERP", "scores": [15, 22, 21, 59]},
        {"question": "What is a transaction code?", "scores": [14, 23, 20, 52]},
        {"question": "Explain MRP", "scores": [11, 25, 22, 60]},
        {"question": "What is S/4HANA?", "scores": [13, 21, 23, 50]},
    ]

    data_to_fill = {
        "Employee Name": "Aditya Sarkale",
        "Evaluator Name": "Sunil Bidla",
        "Dept./ Section": "IT",
        "Current Skill level": "Expert",
        "Rev Date :": "2025-07-10",
        "Prepared  By:-": "VijayaLaxmi",
        "Approved By:-": "Avinash Chormale",
        "Questionaries": questions
    }

    fill_excel_template(data_to_fill, "Aditya_Sarkale_SkillMatrix.xlsx")
